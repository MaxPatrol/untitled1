from openpyxl import *
from openpyxl.styles import Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pprint import pprint

from tkinter import filedialog
from tkinter import *

inotab = "ETOPAHKCBeopakc"
outtab = "ЕТОРАНКСВеоракс"
trantab = str.maketrans(inotab, outtab)

#Boxs -  [ ["NameBox1",[ ["NameRack1",[ListModules]], ["NameRack2",[ListModules]]...]], ...]
#SelectedFiles - список имен файлов, выбранных через диалоговое окно



root = Tk()

SelectedFiles = filedialog.askopenfilenames(initialdir = "",title = "Выбор файлов серийников модулей",filetypes = (("xlsx","*.xlsx"),("all files","*.*")))
print ("Выбраны файлы:")
for i in SelectedFiles:
    print(i)
print ("")

Boxs =[]
for file in SelectedFiles:
    print("Обрабатывается файл {0} ...".format(file))
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]

    print("Поиск имени шкафа ...")
    NameBox=''
    Box = []
    #Поиск имени в первой строке
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            temp = cell.value
            #print(temp)
            if temp != None:
                if re.search('[a-zA-Z]', temp):
                    print("АХТУНГ!! {0}, {1}".format(temp, re.findall('[a-zA-Z]', temp)))
                    print("АХТУНГ!!. Исправление...")
                    temp = temp.translate(trantab)
                    if re.search('[a-zA-Z]', temp):
                        print("АХТУНГ!!. Исправление...Провал!")
                    else:
                        print("АХТУНГ!!. Исправлено!")
                if "ПТ" in temp:
                    NameBox=temp
                    Box.append(NameBox)
                    print("Найдено имя шкафа: {0}".format(temp))
    print("")

    print("Поиск модулей...")
    #TODO Упростить до поиска по шаблону модуля
    pairsPlaceType = []
    for row in ws.rows:
        placeModule = row[1].value
        typeModule = row[2].value
        if placeModule != None and len(placeModule)>3 and typeModule != None and len(typeModule) > 2:
            #print("Обрабатываем строку")


            if placeModule[0] == 'A' and (placeModule[2] == '.' or placeModule[3] == '.') and re.search('(ST|PP|CU|DI|AI|DO|CP) \d{2} \d{3}',typeModule):
                pairsPlaceType.append([placeModule, typeModule])
                #print ([placeModule,typeModule])

    print("Поиск корзин...")

    NameRacks = []

    Racks =[]
    for item in pairsPlaceType:
        if (item[0][:2] in NameRacks)== FALSE:
            NameRacks.append(item[0][:2])
    #print(NameRacks)

    for item in NameRacks:
        modules=[]
        Rack =[]
        Rack.append(item)
        for item2 in pairsPlaceType:
            if (item == item2[0][:2]):
                modules.append(item2[1])
        print("Корзина {0}: {1}".format(item,modules))
        Rack.append(modules)
        Racks.append(Rack)

    print("")
    Box.append(Racks)
    Boxs.append(Box)

Boxs = sorted(Boxs, key=lambda a: a[0])
print("Результаты: ")
for item in Boxs:
    print("Шкаф: {0}".format(item[0]))
    for item2 in item[1]:
        print("Корзина: {0}".format(item2[0]))

        for item3 in item2[1]:
            print(item3)
        print("")
    print("")



##############################################
#Запись результатов в файл
##############################################


# Cоздаем файл для записи результатов
wb = Workbook()

dest_filename = 'empty_book.xlsx'
ws = wb.active
ws.title = 'Модули'

#Шаблон для КЦ(обязательный)

#Серая заливка неиспользуемых ячеек

a= PatternFill("solid", fgColor="808080")
for i in range(204):
    tc = ws.cell(row=i + 1, column=1)
    tc.fill = a
    tc = ws.cell(row=i + 1, column=31)
    tc.fill = a
    tc = ws.cell(row=i + 1, column=32)
    tc.fill = a
    tc = ws.cell(row=i + 1, column=33)
    tc.fill = a

for i in range(30):
    for i2 in range(5, 204,6):
        tc = ws.cell(row=i2, column=i+1)
        tc.fill = a
        tc = ws.cell(row=i2+1, column=i+1)
        tc.fill = a

# Установка ширины столбцов
ws.column_dimensions[get_column_letter(1)].width = 7
ws.column_dimensions[get_column_letter(2)].width = 14
for i in range(29):
    ws.column_dimensions[get_column_letter(3+i)].width = 8

# Установка высоты строк
for i in range(1, 204,6):
    ws.row_dimensions[i].height = 11
    ws.row_dimensions[i+1].height = 73
    ws.row_dimensions[i+2].height = 9
    ws.row_dimensions[i+3].height = 12
    ws.row_dimensions[i+4].height = 12
    ws.row_dimensions[i+5].height = 12

# Расставляем объедененные ячейки со стилями
thin = Side(border_style='medium', color="000000")
all_border_cell = Border(top=thin, left=thin, right=thin, bottom=thin)
centred_cell_style = Alignment(horizontal="center", vertical="center")

a = [['A1',1],['A2',7],['A3',13],['A4',19],['A5',25],['A6',31]]
for i in a:
    tc = ws.cell(i[1],3,i[0])
    tc.border = all_border_cell;
    tc.alignment = centred_cell_style
    ws.merge_cells(start_row=i[1], start_column=3, end_row=i[1]+3, end_column=3)

a = [['КЦ',1],['КС',13],['КК',25]]
for i in a:
    tc = ws.cell(i[1],30,i[0])
    tc.border = all_border_cell;
    tc.alignment = centred_cell_style
    ws.merge_cells(start_row=i[1], start_column=30, end_row=i[1]+9, end_column=30)


#################################
# Заполняем шкаф КЦ данными
#################################
i1=1
for item in Boxs[0][1]:
    print (item)
    ws.cell(i1  , 2, Boxs[0][0])
    ws.cell(i1  , 3,item[0])
    ws.cell(i1+2  , 2,"Начальный адрес")
    ws.cell(i1+3  , 2,"Конечный адрес")
    i3=4
    for item2 in item[1]:
        print ("R500 "+item2)
        ws.cell(i1+1, i3, "R500 "+item2)
        i3 = i3 +1
    i1 = i1 + 6
print("")
print("")
#################################
# Заполняем остальные шкафы
#################################
i1=37
for item0 in Boxs[1:]:
    for item in item0[1]:
        print (item)
        ws.cell(i1  , 2, item0[0])
        ws.cell(i1+2  , 2,"Начальный адрес")
        ws.cell(i1+3  , 2,"Конечный адрес")
        tc = ws.cell(i1,3,item[0])
        tc.border = all_border_cell;
        tc.alignment = centred_cell_style
        ws.merge_cells(start_row=i1, start_column=3, end_row=i1+3, end_column=3)

        i3=4
        for item2 in item[1]:
            print ("R500 "+item2)
            ws.cell(i1+1, i3, "R500 "+item2)
            i3 = i3 +1
        i1 = i1 + 6

#################################
#Обработка ТБ
################################
AI=[]
DI=[]
DO=[]
SelectedFiles = filedialog.askopenfilenames(initialdir = "",title = "Выбор файлов ТБ",filetypes = (("xlsx","*.xlsx"),("all files","*.*")))
print ("Выбраны файлы:")
for i in SelectedFiles:
    print(i)
print ("")

rawTBs=[]
TBs =[]
for file in SelectedFiles:
    rawTBs =[]
    print("Обрабатывается файл {0} ...".format(file))
    wb = load_workbook(file)
    ws = wb[wb.sheetnames[0]]


    print("Поиск имени шкафа...")
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            temp = cell.value
            #print(temp)
            if temp != None:
                if re.search('[a-zA-Z]', temp):
                    print("АХТУНГ!! {0}, {1}".format(temp, re.findall('[a-zA-Z]', temp)))
                    print("АХТУНГ!!. Исправление...")
                    temp = temp.translate(trantab)
                    if re.search('[a-zA-Z]', temp):
                        print("АХТУНГ!!. Исправление...Провал!")
                    else:
                        print("АХТУНГ!!. Исправлено!")
                if "ПТ" in temp:
                    NameBox=temp
                    Box.append(NameBox)
                    print("Найдено имя шкафа: {0}".format(temp))
    print("")


    for row in ws.rows:
        tb =[]
        for item in row[:5]:
            tb.append(item.value)
        rawTBs.append(tb)

    for item in rawTBs:
        if item ==['Идентификатор', 'Наименование сигнала', '№ Корзины', '№ модуля в корзине', '№ канала в модуле']:
            print ("Поддвержден патерн ТБ")



    for item in rawTBs:
        if item ==[1,2,3,4,5]:
            print ("Определно начало данных {0}".format(rawTBs.index([1, 2, 3, 4, 5])))

    for i in range(rawTBs.index([1, 2, 3, 4, 5]) + 1):
        rawTBs.pop(0)

    a=[]
    for item in rawTBs:
        if item not in a:
            a.append(item)
    rawTBs =a
    rawTBs.remove([None, None, None, None, None])

    # for item in rawTBs:
    #     print (item)

    print(len(rawTBs))

    print("Поиск AI...")

    for item in rawTBs:
        if (item[4][:2] == "AI") and (item[1] != None) and (item[1].lower() != "резерв") :
            AI.append(item)

    print("Поиск DO...")

    for item in rawTBs:
        if (item[4][:2] == "DO") and (item[1] != None) and (item[1].lower() != "резерв") :
            DO.append(item)





DO = sorted(DO,key=lambda a:a[1])

AI = sorted(AI,key=lambda a:a[1])

for item in AI:
    print(item)
print("")

for item in DO:
    print(item)

print ("")
print ("Поиск защищаемых зон...")
print ("")
# Правило 1.
# Находим все DO с подстроками "пожар в " и "сигнал в"
# выделяем в отдельный список
Rule1=[]
for item in DO:
    if ("сигнал в" in item[1].lower() and ("пожар в" in item[1].lower() or "пожар на" in item[1].lower())):
        Rule1.append(item[1])
for item in Rule1:
    print (item)

# Правило 2.
# Вырезаем проектное обозначение защищаемых объектов, нпример, (003.2)
# т.е. подстроку любое количество символов между скобок, включая скобки
#TODO любое количество цифр или точка
Rule2=[]
for item in Rule1:
    a=re.sub(' \(.*?\)',"", item,1)
    a=re.sub('Пожар в ',"", a,1)
    a=re.sub('Пожар на ',"", a,1)
    Rule2.append(a)

for item in Rule2:
    print (item)

# Правило 3.
# Если есть точка, удаляем все что после нее, и точку тоже

Rule3 = []
for item in Rule2:
    if "." in item:
        b = item.find('.')
    else:
        b = None
    a = item[:b]
    Rule3.append(a)

for item in Rule3:
    print (item)

# Правило 4.
# Убираем дубликаты

Rule4 = []
for item in Rule3:
    if item not in Rule4:
        Rule4.append(item)

for item in Rule4:
    print (item)





#pprint(TBs)

ws2 = wb.create_sheet(title="Аналоги")

print('Запись файла')
wb.save(filename=dest_filename)
